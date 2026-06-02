import os
from datetime import datetime
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from models import AnalysisResult, ProductItem


def export_to_excel(
    output_path: str,
    keyword: str,
    products: List[ProductItem],
    analysis: AnalysisResult,
    wordcloud_path: str,
) -> str:
    output_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _build_raw_data_sheet(wb, keyword, products)
    _build_rank_sheet(wb, "Title Keywords", analysis.title_keywords)
    _build_rank_sheet(wb, "Selling Keywords", analysis.selling_keywords)
    _build_spec_sheet(wb, analysis)
    _build_recommendation_sheet(wb, analysis)
    _build_wordcloud_sheet(wb, wordcloud_path)

    wb.save(output_path)
    return output_path


def _build_raw_data_sheet(wb: Workbook, keyword: str, products: List[ProductItem]) -> None:
    ws = wb.create_sheet("Product Data")
    ws.append(["Search Keyword", keyword])
    ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["No.", "Title", "Price", "Reviews", "Product Link"])

    for idx, item in enumerate(products, start=1):
        ws.append([idx, item.title, item.price, item.reviews, item.link])

    _style_header(ws, row_idx=4, col_count=5)
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 60

    for row in ws.iter_rows(min_row=5, max_row=4 + len(products), min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=5, max_row=4 + len(products), min_col=5, max_col=5):
        for cell in row:
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def _build_rank_sheet(wb: Workbook, sheet_name: str, rows: List[Tuple[str, int]]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(["Rank", "Keyword", "Count"])
    for idx, (word, count) in enumerate(rows, start=1):
        ws.append([idx, word, count])

    _style_header(ws, row_idx=1, col_count=3)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12


def _build_spec_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Hot Specs")
    ws.append(["Category", "Value", "Count"])

    _append_section(ws, "Hot Specs", analysis.hot_specs)
    _append_section(ws, "Hot Lengths", analysis.hot_lengths)
    _append_section(ws, "Hot LED Counts", analysis.hot_led_counts)

    _style_header(ws, row_idx=1, col_count=3)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12


def _build_recommendation_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Recommendations")
    ws.append(["Type", "Content"])
    _style_header(ws, row_idx=1, col_count=2)

    ws.append(["TEMU Title Suggestions", ""])
    for title in analysis.title_recommendations:
        ws.append(["", title])

    ws.append([])
    ws.append(["Differentiation Suggestions", ""])
    for suggestion in analysis.differentiation_suggestions:
        ws.append(["", suggestion])

    ws.append([])
    ws.append(["AI Image Prompts", ""])
    for prompt in analysis.ai_prompts:
        ws.append(["", prompt])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _build_wordcloud_sheet(wb: Workbook, wordcloud_path: str) -> None:
    ws = wb.create_sheet("Word Cloud")
    ws.append(["Keyword Word Cloud"])
    _style_header(ws, row_idx=1, col_count=1)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 24

    if os.path.exists(wordcloud_path):
        img = XLImage(wordcloud_path)
        img.width = 1000
        img.height = 620
        ws.add_image(img, "A3")
    else:
        ws["A3"] = "Word cloud image was not generated."


def _append_section(ws, section_name: str, rows: List[Tuple[str, int]]) -> None:
    if not rows:
        ws.append([section_name, "No data", 0])
        return
    for value, count in rows:
        ws.append([section_name, value, count])


def _style_header(ws, row_idx: int, col_count: int) -> None:
    fill = PatternFill(fill_type="solid", start_color="0D47A1", end_color="0D47A1")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
